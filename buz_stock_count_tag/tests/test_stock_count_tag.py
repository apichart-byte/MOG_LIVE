import base64
import io

import openpyxl
from odoo.tests import TransactionCase, tagged


@tagged("-at_install", "post_install")
class TestStockCountTag(TransactionCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Tag = cls.env["buz.stock.count.tag"]
        cls.Line = cls.env["buz.stock.count.tag.line"]

    def test_tag_sequence_name(self):
        tag = self.Tag.create({})
        self.assertTrue(tag.name.startswith("TAG/"))

    def test_tag_and_page_count(self):
        tag = self.Tag.create({})
        self.Line.create(
            {
                "tag_id": tag.id,
                "product_code": "TEST001",
                "product_name": "Test Product",
                "warehouse": "WH1",
                "quantity": 10.0,
            }
        )
        self.assertEqual(tag.tag_count, 1)
        self.assertEqual(tag.page_count, 1)

    def test_generate_excel_structure(self):
        tag = self.Tag.create({})
        for i in range(3):
            self.Line.create(
                {
                    "tag_id": tag.id,
                    "sequence": i,
                    "product_code": f"PROD{i:03d}",
                    "product_name": f"Product {i}",
                    "warehouse": "WH1",
                    "quantity": float(i + 1),
                }
            )
        tag.state = "imported"
        tag.sort_by = "import_order"

        result = tag.action_generate_excel()
        self.assertEqual(result["type"], "ir.actions.act_url")
        self.assertTrue(tag.generated_file_id)

        file_bytes = base64.b64decode(tag.generated_file_id.datas)
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active

        self.assertEqual(len(ws.row_breaks.brk), 3)

        # first tag (rows 1-24)
        self.assertEqual(ws["C5"].value, "PROD000")
        self.assertEqual(ws["C16"].value, 1.0)
        self.assertEqual(ws["J5"].value, ws["C5"].value)
        self.assertEqual(ws["J16"].value, ws["C16"].value)

        # second tag (rows 25-48)
        self.assertEqual(ws["C29"].value, "PROD001")
        self.assertEqual(ws["C40"].value, 2.0)
        self.assertEqual(ws["J29"].value, ws["C29"].value)
