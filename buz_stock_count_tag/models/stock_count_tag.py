import base64
import copy
import io

from odoo import _, api, fields, models
from odoo.exceptions import UserError
from odoo.modules.module import get_module_resource

try:
    import openpyxl
    from openpyxl.styles import Alignment
    from openpyxl.worksheet.pagebreak import Break
except ImportError:
    openpyxl = None
    Alignment = None
    Break = None

TEMPLATE_BLOCK_HEIGHT = 24
# row offsets (within one 24-row block) of each field's row
F_ROW = {
    "doc_no": 2,
    "date": 3,
    "product_code": 5,
    "product_name": 6,
    "product_status": 8,
    "warehouse": 10,
    "layout": 12,
    "lac": 14,
    "quantity": 16,
}
# The template's label cell for "doc_no"/"date" sits in column D (ต้นฉบับ)
# / K (สำเนา), so their value must land in the wide column immediately to
# its right (E / L) to sit flush next to the label, matching the physical
# form (~/Desktop/ตัวอย่าง TAG.jpg). Every other field's label sits in
# column A / H, so its value lands in the next wide column (C / J).
# Column F / M only ever held a leftover font/border style, never the
# actual fill-in position — confirmed against the reference photo.
VALUE_COL = {
    "doc_no": (5, 12),  # E, L
    "date": (5, 12),
    "product_code": (3, 10),  # C, J
    "product_name": (3, 10),
    "product_status": (3, 10),
    "warehouse": (3, 10),
    "layout": (3, 10),
    "lac": (3, 10),
    "quantity": (3, 10),
}
STYLE_REF_COL = 6  # F: still carries the template's Angsana New 16 font


class StockCountTag(models.Model):
    _name = "buz.stock.count.tag"
    _description = "Stock Count Tag"
    _inherit = ["mail.thread", "mail.activity.mixin"]
    _order = "id desc"

    name = fields.Char(default="New", copy=False, readonly=True)
    date = fields.Date(default=fields.Date.context_today, tracking=True)
    state = fields.Selection(
        [
            ("draft", "Draft"),
            ("imported", "Imported"),
            ("generated", "Generated"),
            ("printed", "Printed"),
        ],
        default="draft",
        tracking=True,
        copy=False,
    )
    sort_by = fields.Selection(
        [
            ("warehouse_layout_lac_product", "Warehouse + Layout + LAC + Product Code"),
            ("product_code", "Product Code"),
            ("warehouse", "Warehouse"),
            ("layout", "Layout"),
            ("lac", "LAC"),
            ("import_order", "Import Order"),
        ],
        default="warehouse_layout_lac_product",
        required=True,
    )
    line_ids = fields.One2many("buz.stock.count.tag.line", "tag_id", string="Lines")
    tag_count = fields.Integer(compute="_compute_tag_count")
    page_count = fields.Integer(compute="_compute_tag_count")
    generated_file_id = fields.Many2one("ir.attachment", readonly=True, copy=False)

    @api.depends("line_ids")
    def _compute_tag_count(self):
        for rec in self:
            rec.tag_count = len(rec.line_ids)
            rec.page_count = len(rec.line_ids)

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if not vals.get("name") or vals["name"] == "New":
                vals["name"] = self.env["ir.sequence"].next_by_code(
                    "buz.stock.count.tag"
                ) or "New"
        return super().create(vals_list)

    def action_printed(self):
        self.write({"state": "printed"})

    def action_open_import_wizard(self):
        self.ensure_one()
        return {
            "type": "ir.actions.act_window",
            "name": _("Import Stock Count Tags"),
            "res_model": "buz.stock.count.tag.import.wizard",
            "view_mode": "form",
            "target": "new",
            "context": {"default_tag_id": self.id},
        }

    def action_open_pdf_export_wizard(self):
        self.ensure_one()
        return {
            "type": "ir.actions.act_window",
            "name": _("Export Stock Count Tag PDF"),
            "res_model": "buz.stock.count.tag.pdf.export.wizard",
            "view_mode": "form",
            "target": "new",
            "context": {"default_tag_id": self.id},
        }

    def action_download_template(self):
        self.ensure_one()
        if not openpyxl:
            raise UserError(_("Python module 'openpyxl' is required."))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import"
        headers = [
            "doc_no",
            "product_code",
            "product_name",
            "product_status",
            "warehouse",
            "layout",
            "lac",
            "quantity",
        ]
        ws.append(headers)
        ws.append(["TAG/2026/00027", "SP00010101", "SP10T/1 ตัวเก็บประจุ", "A", "คลังบริการเทคนิค", "ห้องบริการเทคนิค", "SV-1", 3])
        ws.append(["TAG/2026/00027", "SP00010102", "Motor Pump", "A", "คลังบริการเทคนิค", "ห้องบริการเทคนิค", "SV-2", 5])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        attachment = self.env["ir.attachment"].create(
            {
                "name": "stock_count_tag_import_template.xlsx",
                "datas": base64.b64encode(buf.read()),
                "res_model": self._name,
                "res_id": self.id,
                "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }
        )
        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }

    def _sorted_lines(self):
        self.ensure_one()
        lines = self.line_ids
        key_map = {
            "warehouse_layout_lac_product": lambda l: (
                l.warehouse or "",
                l.layout or "",
                l.lac or "",
                l.product_code or "",
            ),
            "product_code": lambda l: l.product_code or "",
            "warehouse": lambda l: l.warehouse or "",
            "layout": lambda l: l.layout or "",
            "lac": lambda l: l.lac or "",
        }
        if self.sort_by == "import_order":
            return lines.sorted(key=lambda l: (l.sequence, l.id))
        return lines.sorted(key=key_map[self.sort_by])

    def _get_pdf_lines(self, doc_no_from=None, doc_no_to=None):
        self.ensure_one()
        lines = self._sorted_lines()
        if doc_no_from and doc_no_to:
            lines = lines.filtered(
                lambda l: doc_no_from <= (l.doc_no or "") <= doc_no_to
            )
        return lines

    @staticmethod
    def _strip_style_from_blank_cells(ws):
        """Drop the style on every value-less cell in the used range.

        openpyxl always serializes a cell that has any style applied, even
        when its value is None/"" — it writes ``<c t="n"/>`` (or an empty
        ``<c t="inlineStr"/>``) with no ``<v>``/``<is>`` child, which Excel
        treats as corrupt and refuses to open. This affects both the
        untouched template cells (font-formatted-but-empty spacer cells)
        and any cell this method wrote an empty value into. Clearing
        ``_style`` makes ``has_style`` False so the writer skips the cell
        entirely, matching how a cell that was never touched serializes.
        """
        for row in ws.iter_rows():
            for cell in row:
                if cell.value in (None, "") and cell.has_style:
                    cell._style = None

    @staticmethod
    def _copy_cell_style(src_cell, dst_cell):
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format

    def action_generate_excel(self):
        self.ensure_one()
        if not openpyxl:
            raise UserError(_("Python module 'openpyxl' is required."))
        if self.state not in ("imported", "generated", "printed"):
            raise UserError(_("Tag must be imported before generating the Excel file."))
        if not self.line_ids:
            raise UserError(_("There are no lines to generate."))

        template_path = get_module_resource(
            "buz_stock_count_tag", "templates", "stock_count_tag_template.xlsx"
        )
        if not template_path:
            raise UserError(_("Template file 'stock_count_tag_template.xlsx' not found."))

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # The template's docProps/core.xml has an empty <dc:creator>/
        # <cp:lastModifiedBy>. openpyxl drops empty core properties on
        # save instead of writing them blank, which leaves docProps/core.xml
        # inconsistent with the OPC package in a way that macOS/Numbers
        # hard-rejects ("file can't be opened") even though the file is a
        # well-formed zip and every other part is spec-valid. Setting these
        # explicitly makes openpyxl emit them non-empty.
        wb.properties.creator = "Mogen Co., Ltd."
        wb.properties.lastModifiedBy = "Mogen Co., Ltd."

        # "Fit to page" scaling makes Excel IGNORE manual row breaks
        # (ws.row_breaks below) — it recomputes its own page boundaries to
        # squeeze everything into the fitToHeight page count instead. Use a
        # fixed 100% scale so each manual break actually starts a new page.
        ws.sheet_properties.pageSetUpPr.fitToPage = False
        ws.page_setup.scale = 100
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        lines = self._sorted_lines()

        for idx, line in enumerate(lines):
            row_offset = idx * TEMPLATE_BLOCK_HEIGHT

            if idx > 0:
                for src_row in range(1, TEMPLATE_BLOCK_HEIGHT + 1):
                    dst_row = src_row + row_offset
                    if src_row in ws.row_dimensions:
                        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
                    for col in range(1, 13):
                        src_cell = ws.cell(row=src_row, column=col)
                        dst_cell = ws.cell(row=dst_row, column=col)
                        dst_cell.value = src_cell.value
                        # Copy style even for blank cells (e.g. column F,
                        # which only carries a font used later as this
                        # page's style reference — skipping it here left
                        # every page after the first with a default-size
                        # font). A blank-but-styled cell would otherwise
                        # serialize as invalid OOXML, but
                        # _strip_style_from_blank_cells() below cleans up
                        # every such cell once, after all pages are built.
                        self._copy_cell_style(src_cell, dst_cell)

            values = {
                # The printed TAG number is per-line, driven by the
                # imported Excel "doc_no" column; fall back to the header's
                # document number for lines added/edited manually without one.
                "doc_no": line.doc_no or self.name,
                "date": self.date,
                "product_code": line.product_code,
                "product_name": line.product_name,
                "product_status": line.product_status,
                "warehouse": line.warehouse,
                "layout": line.layout,
                "lac": line.lac,
                "quantity": line.quantity,
            }
            for field_key, row_in_block in F_ROW.items():
                actual_row = row_in_block + row_offset
                style_ref = ws.cell(row=actual_row, column=STYLE_REF_COL)
                left_col, right_col = VALUE_COL[field_key]
                left_cell = ws.cell(row=actual_row, column=left_col)
                left_cell.font = copy.copy(style_ref.font)
                value = values[field_key]
                if field_key == "quantity":
                    left_cell.value = float(value or 0.0)
                    left_cell.number_format = "0.00"
                    # Leave alignment on General: numbers default to
                    # right-aligned, which sits the value flush against the
                    # "ชิ้น" unit label immediately to its right, matching
                    # the physical form.
                elif field_key == "date":
                    left_cell.value = value
                    left_cell.number_format = "yyyy-mm-dd"
                    left_cell.alignment = Alignment(horizontal="left")
                else:
                    # None (not "") for blanks: an empty string is still a
                    # non-None value, so openpyxl's own "skip blank cells"
                    # check never fires for it and it gets serialized as an
                    # empty <c t="inlineStr"/> with no <is> child.
                    left_cell.value = value or None
                    left_cell.alignment = Alignment(horizontal="left")

                right_cell = ws.cell(row=actual_row, column=right_col)
                self._copy_cell_style(left_cell, right_cell)
                right_cell.value = left_cell.value

            block_last_row = row_offset + TEMPLATE_BLOCK_HEIGHT
            ws.row_breaks.append(Break(id=block_last_row))

        self._strip_style_from_blank_cells(ws)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        file_name = f"{self.name.replace('/', '_')}_stock_count_tag.xlsx"
        attachment = self.env["ir.attachment"].create(
            {
                "name": file_name,
                "datas": base64.b64encode(buf.read()),
                "res_model": self._name,
                "res_id": self.id,
                "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }
        )
        self.write({"generated_file_id": attachment.id, "state": "generated"})

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }

    def action_generate_pdf(self, doc_no_from=None, doc_no_to=None):
        self.ensure_one()
        if self.state not in ("imported", "generated", "printed"):
            raise UserError(_("Tag must be imported before generating the PDF file."))
        if not self.line_ids:
            raise UserError(_("There are no lines to generate."))
        if doc_no_from and doc_no_to and not self._get_pdf_lines(doc_no_from, doc_no_to):
            raise UserError(_("No lines found in that Tag No. range."))

        # Large tag batches (1000+ pages) time out or crash the browser's
        # inline PDF preview (ir.actions.report's default qweb-pdf action).
        # Render server-side instead and hand back a plain attachment
        # download, same pattern as action_generate_excel above.
        report = self.env["ir.actions.report"]._get_report_from_name(
            "buz_stock_count_tag.report_stock_count_tag_document"
        )
        data = {"doc_no_from": doc_no_from, "doc_no_to": doc_no_to}
        pdf_content, _report_type = report._render_qweb_pdf(
            "buz_stock_count_tag.report_stock_count_tag_document", self.ids, data=data
        )

        file_name = f"{self.name.replace('/', '_')}_stock_count_tag.pdf"
        if doc_no_from and doc_no_to:
            suffix = f"{doc_no_from}-{doc_no_to}".replace("/", "_")
            file_name = f"{self.name.replace('/', '_')}_{suffix}_stock_count_tag.pdf"
        attachment = self.env["ir.attachment"].create(
            {
                "name": file_name,
                "datas": base64.b64encode(pdf_content),
                "res_model": self._name,
                "res_id": self.id,
                "mimetype": "application/pdf",
            }
        )

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }
